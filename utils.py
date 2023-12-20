import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

# Functions from our previous implementation
def resize_image_to_columns(img, num_columns):
    standard_column_width_px = 64
    new_width = standard_column_width_px * num_columns
    ratio = new_width / img.width
    new_height = img.height * ratio
    img.width = new_width
    img.height = new_height

def append_df_and_image_to_new_ws(wb, df, img_path, sheet_name):
    """
    Append a dataframe and its corresponding image to a new worksheet in the workbook.
    """
    ws = wb.create_sheet(title=sheet_name)

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    img = Image(img_path)
    resize_image_to_columns(img, len(df.columns))

    start_col = len(df.columns) + 3
    col_letter = openpyxl.utils.get_column_letter(start_col)

    ws.add_image(img, "{}1".format(col_letter))